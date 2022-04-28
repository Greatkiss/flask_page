#owner page
#https://owner.tabelog.com/
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl as xl
import shutil
import time
import os
from Scraping_gnavi import csv_to_xlsx
import zipfile

e_list = []
folder_path = "./Scraping_gnavi/results"

class get_table():
    @staticmethod
    def golden(name, id, pwd, date):
        session = requests.session()
        try:
            # セッションを開始
            url_login = "https://ssl.tabelog.com/owner_account/login/"
            response = session.get(url_login)
            res = session.get(url_login)
            soup = BeautifulSoup(res.text,"html.parser")
            auth_token = soup.find('input', {'name': 'authenticity_token'}).get('value')
            response_cookie = response.cookies
            # ログイン
            login_info = {
                "utf8":"✓",
                "authenticity_token":auth_token,
                "login_id":id,  #"login_id":id
                "password":pwd, #"password":pwd
                "login":"1",
                "Target":""
            }
            # action
            res = session.post(url_login, data=login_info, cookies=response_cookie)
            res.raise_for_status() # エラーならここで例外を発生させる
            soup_1 = BeautifulSoup(res.text,"html.parser")
            mp_link = soup_1.select('#wrapper > div.op-rst-navi.op-rst-navi--owner > ul > li:nth-child(1)')[0].select('a')[0].get('href')
            res = session.get(mp_link)
            res.raise_for_status()
            soup_2 = BeautifulSoup(res.text,"html.parser")
            golden_time_link = soup_2.select('#wrapper > div.ol-container > div > div > div.ol-column2__side > div > div.op-side-nav > ul > li:nth-child(11)')[0].select('a')[0].get('href')
            #golden_time
            res = session.get(golden_time_link)
            res.raise_for_status()
            #tableの取得
            raw_table=pd.DataFrame()
            table = pd.DataFrame()
            raw_table = pd.read_html(res.text)
            #golden_timeのtable取得
            table = raw_table[2]
            table.rename(columns={'曜日（成約数）  ゴールデンタイム':'date'}, inplace=True)
            table.rename(columns={'曜日（成約数）  ゴールデンタイム.1':'time'}, inplace=True)
            table.rename(columns={'コール':'call'}, inplace=True)
            table.rename(columns={'ネット予約':'online'}, inplace=True)
            table.set_index(['date'], inplace=True)

            #selected-timeを取得
            soup_table = BeautifulSoup(res.text,"html.parser")
            selected_time_raw = soup_table.select('#wrapper > div.ol-container > div > div > div.ol-column2__main > div > div.weekly-intro-effect > table > tbody')[0].select('tr')
            selected_time_list = []
            for html_each_time in selected_time_raw:
                th_name = html_each_time.find('td', class_='weekly-intro-effect__cell is-selected')
                if th_name is None:
                    selected_time_list.append(0)
                else:
                    selected_time_list.append(1)
            table['selection']= selected_time_list

            table.to_csv(folder_path + "/{}_{}.csv".format(name, date), encoding='shift-jis')
            time.sleep(1)
            #logout
            logout_link = soup_table.select('#owner-headline > ul > li:nth-child(3) > a')[0].get('href')
            res = session.get(logout_link)
            soup_out = BeautifulSoup(res.text,"html.parser")
            relogin_link = soup_out.select('#js-body > div.ol-container.ol-container--mat > div.ol-contents.ol-contents--fixed > div > p:nth-child(2) > a')[0].get('href')
            session.get(relogin_link)
            session.close()
        except:
            print("error at {}".format(name))
            e_list.append([name, id, pwd])
            session.close()
    def get_idpwd(n, path):
        niplist = pd.DataFrame()
        for i in range(n):
            book = xl.load_workbook(path,data_only=True)
            sheet = book["nip"]
            name = str(sheet.cell(row=i+1,column=1).value)  #forは0から，xlsxの1行目はindex，よってi+2でループ
            id = str(sheet.cell(row=i+1,column=2).value)
            pwd = str(sheet.cell(row=i+1,column=3).value)
            nip = pd.DataFrame([[id, pwd]],columns=['id', 'pwd'], index=[name])
            niplist = pd.concat([niplist, nip])
        shutil.rmtree("uploads")
        return niplist

def main(num, path, weekdate):
    #IDとパスワードを"./idpass.xlsx"から取得し，pandas dataframeに格納
    niplist = get_table.get_idpwd(num,path) #get_idpwdの引数は読み込む法人数
    #niplistからidとpassを一行ずつ取得して，ログイン・スクレイプ・店舗ごとにcsvで保存
    os.mkdir(folder_path)
    for i in range(len(niplist)):
        print("log in to {}".format(niplist.index[i]))
        get_table.golden(niplist.index[i],niplist.iat[i,0], niplist.iat[i,1], weekdate)
    #生成したcsvを一つにまとめて保存し、ディレクトリ名を変更
    csv_to_xlsx.excelize(folder_path)
    new_path = "./Scraping_gnavi/results_{}".format(weekdate)
    os.rename(folder_path, new_path)
    #errorファイルの作成・保存
    pderror=pd.DataFrame(e_list)
    pderror.to_csv(new_path+"/errorlist.csv",encoding='shift-jis', index=False)
    #results_{}フォルダのzip化
    shutil.make_archive('results', 'zip', root_dir=new_path)
    shutil.rmtree(new_path)

if __name__ == "__main__":
    main(1,"./uploads/idpass.xlsx")